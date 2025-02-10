#!/usr/bin/env python3
import pandas as pd
import argparse
from openpyxl import load_workbook
from collections import namedtuple
from dataclasses import dataclass

# TODO:
# - add a way to specify the starting position for a ticker (e.g. if you already had some shares before the export)
# - do the fifo calculation
# - keep track of conversion gains/losses

# negative values decrement the cash on balance
# so buy trades are negative
# new card cost is negative
# currency conversion fee is negative
#
# positive values increment the cash on balance
# so sell trades are positive
# interest on cash is positive
# lending interest is positive
# deposit is positive
# dividend is positive
# dividend manufactured payment is positive


class trading_export_sorter:
    def __init__(self, input_file, debug=False):
        self.debug = debug

        self.df = pd.read_csv(input_file)
        self.debug_print(self.df.columns)
        self.debug_print(f"Actions in the file: {self.df['Action'].unique()}")
        # invert values of currency conversion fees
        self.df['Currency conversion fee'] *= -1
        # invert values of buy trades
        buy_columns_to_invert = ['Total', 'No. of shares']
        self.df.loc[self.df['Action'].str.lower().str.contains('buy'), buy_columns_to_invert] *= -1
        self.buy_and_sell = self.df.loc[self.df['Action'].isin(["Limit buy", "Limit sell", "Market buy", "Market sell"])].groupby('Ticker')
       
    def calculate_fifo_for_ticker(self, ticker):
        Trade = namedtuple('Trade', ['shares', 'price'])
        @dataclass
        class Trade:
            shares: float
            price: float

        trades = []
        result = 0

        for _, row in ticker.iterrows():
            incomming_trade_action = row['Action'].lower()
            incomming_trade = Trade(shares=row['No. of shares'], price=row['Price / share'])
            self.debug_print(f"----------------------   ")
            for i, trade in enumerate(trades):
                self.debug_print(f"  Trade {i+1:3d}: {trade.shares:>10.2f} shares @ ${trade.price:>4.2f}")
            self.debug_print(f"current result: {result}")
            self.debug_print(f"----------------------   ")
            self.debug_print(f"incomming row: {row['Action']} {row['No. of shares']:.2f} {row['Price / share']:.2f}")

            if "buy" in incomming_trade_action:
                trades.append(incomming_trade)
            elif "sell" in incomming_trade_action:
                for trade in trades:
                    if abs(trade.shares) >= abs(incomming_trade.shares):
                        trade.shares += incomming_trade.shares
                        result += (incomming_trade.price - trade.price) * abs(incomming_trade.shares)
                        break
                    else:
                        incomming_trade.shares += trade.shares
                        result += (incomming_trade.price - trade.price) * abs(trade.shares)
                        trade.shares = 0
        resulting_shares = sum(trade.shares for trade in trades)
        self.debug_print(f"result: {row["Ticker"]} {resulting_shares=} {result=}")
        return result

    def do_work(self, output_file):
        result_sum = 0
        fifo_result_sum = 0
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for ticker_name in self.buy_and_sell.groups.keys():
                ticker = self.buy_and_sell.get_group(ticker_name)
                result = self.calculate_fifo_for_ticker(ticker)
                fifo_result_sum += result
                # add a total row
                columns_to_sum = ['Total', 'No. of shares', 'Result', 'Currency conversion fee']
                sums = ticker[columns_to_sum].sum(numeric_only=True)
                sums.name = 'Total'
                ticker = pd.concat([ticker, sums.to_frame().T])
                result_sum += sums['Result']
                # add a fifo row
                fifo_row = pd.DataFrame({'Result': result}, index=['Fifo'])
                ticker = pd.concat([ticker, fifo_row])
                
                self.debug_print(f"exporting ticker: {ticker_name}")
                ticker.to_excel(writer, sheet_name=ticker_name, index=True)
                #break

            def get_sum(column_name):
                return {column_name:self.df.loc[self.df['Action'] == column_name]['Total'].sum()}
            
            data = dict()
            data.update(get_sum('Interest on cash'))
            data.update(get_sum('Lending interest'))
            data.update(get_sum('Deposit'))
            data.update(get_sum('Dividend (Dividend)'))
            data.update(get_sum('Dividend (Dividend manufactured payment)'))
            data.update(get_sum('New card cost'))
            data.update({'Currency Conversion Fees': self.df["Currency conversion fee"].sum()})
            data.update({'Total results': result_sum})
            data.update({'Fifo total results': fifo_result_sum})
            self.debug_print(data)

            df_main = pd.DataFrame(list(data.items()))
            df_main.to_excel(writer, sheet_name='Main', index=False, header=False)

        self.adjust_column_widths(output_file)
        self.move_main_sheet_to_front(output_file)
        
    def move_main_sheet_to_front(self, output_file):
        workbook = load_workbook(output_file)
        sheet_names = workbook.sheetnames
        if 'Main' in sheet_names and sheet_names[0] != 'Main':
            main_idx = sheet_names.index('Main')
            workbook._sheets.insert(0, workbook._sheets.pop(main_idx))
        workbook.save(output_file)
    
        
    def adjust_column_widths(self, output_file): 
        # Load the workbook to adjust column widths
        workbook = load_workbook(output_file)
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 3
        workbook.save(output_file)
        
    def debug_print(self, *args, **kwargs):
        if self.debug:
            print(*args, **kwargs)



def main():
    parser = argparse.ArgumentParser(description="Process exported data from trading212.")
    parser.add_argument("--output-file", default="output.xlsx", help="Path to the output Excel file")
    parser.add_argument("--debug", action="store_true", help="Enable debug mode")
    parser.add_argument("csv_file", help="Path to the trading212 export CSV file")
    args = parser.parse_args()

    sorter = trading_export_sorter(args.csv_file, debug=args.debug)
    #sorter.calculate_fifo_for_ticker(sorter.buy_and_sell.get_group('3MST'))
    sorter.do_work(args.output_file)

if __name__ == "__main__":
    main()
